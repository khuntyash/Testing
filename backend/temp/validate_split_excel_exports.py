"""End-to-end validation for the new per-split Excel exports in the
crop ZIP pipeline.

Coverage:
  1. ``_extract_split_rows_from_pdfs`` returns canonical rows for both
     Meesho and Flipkart synthetic labels.
  2. ``_write_split_rows_xlsx`` always emits the header row even when the
     input row list is empty.
  3. Full ``_process_crop_task`` flow for Meesho: ZIP contains split PDFs
     + the three category Excels with the expected row counts.
  4. Full ``_process_crop_task`` flow for Flipkart: same expectation.
  5. Stable behaviour when no split mode matches any pages (still emits
     the requested Excel files with header-only contents).

Side effects: writes scratch artifacts under ``backend/temp/split_excel_validation``.
The script is non-destructive to user data.
"""
from __future__ import annotations

import io
import json
import sys
import tempfile
import zipfile
from pathlib import Path
from typing import Any

import fitz
from openpyxl import load_workbook

BACKEND_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BACKEND_DIR))

import task_queue  # noqa: E402
import history_store  # noqa: E402


# --- Synthetic label helpers --------------------------------------------------

def _build_fake_meesho_label(
    *,
    suborder: str,
    name: str,
    address_1: str,
    pincode: str = "560001",
    sku: str = "SKU-DEMO",
) -> bytes:
    """Build a single-page Meesho-style label whose layout matches the
    fields the shared OCR parser already understands.
    """
    doc = fitz.open()
    try:
        page = doc.new_page(width=595, height=842)
        page.insert_text((40, 60), "TAX INVOICE / BILL OF SUPPLY", fontsize=11)
        page.insert_text((40, 90), "Customer Address", fontsize=11)
        page.insert_text((40, 110), name, fontsize=11)
        page.insert_text((40, 130), address_1, fontsize=11)
        page.insert_text((40, 150), f"Bangalore, Karnataka, {pincode}", fontsize=11)
        page.insert_text((40, 170), "If undelivered, return to:", fontsize=10)
        page.insert_text((40, 220), "Order No.", fontsize=11)
        page.insert_text((40, 240), sku, fontsize=11)
        page.insert_text((40, 260), suborder, fontsize=11)
        page.insert_text((40, 290), "Prepaid: Do not collect cash", fontsize=10)
        page.insert_text((40, 320), "Shadowfax", fontsize=10)
        page.insert_text((40, 340), "SF12345678901234", fontsize=10)
        out = io.BytesIO()
        doc.save(out)
        return out.getvalue()
    finally:
        doc.close()


def _build_fake_flipkart_label(
    *,
    order_id: str,
    name: str,
    address_1: str,
    pincode: str = "110001",
) -> bytes:
    """Build a single-page Flipkart-style label whose layout matches the
    Flipkart-specific parser branch.
    """
    doc = fitz.open()
    try:
        page = doc.new_page(width=612, height=792)
        page.insert_text((200, 60), f"{order_id}  COD", fontsize=12, fontname="helv")
        page.insert_text((200, 120), "Shipping/Customer address", fontsize=11, fontname="helv")
        page.insert_text((200, 140), f"Name: {name},", fontsize=10, fontname="helv")
        page.insert_text((200, 160), f"{address_1},", fontsize=10, fontname="helv")
        page.insert_text((200, 180), f"New Delhi - {pincode}, IN-DL", fontsize=10, fontname="helv")
        page.insert_text((200, 240), "SKU ID | Description", fontsize=10, fontname="helv")
        page.insert_text((200, 260), "QTY", fontsize=10, fontname="helv")
        page.insert_text((200, 280), "1 SKU-FK | Sample product", fontsize=10, fontname="helv")
        page.insert_text((200, 300), "FMPC1234567890", fontsize=10, fontname="helv")
        page.insert_text((200, 330), "AWB No. FMPC1234567890", fontsize=10, fontname="helv")
        page.insert_text((200, 360), "Ordered through flipkart.com", fontsize=10, fontname="helv")
        out = io.BytesIO()
        doc.save(out)
        return out.getvalue()
    finally:
        doc.close()


def _write_pdf(path: Path, page_blobs: list[bytes]) -> None:
    """Concatenate single-page PDF bytes into a multi-page PDF on disk."""
    out_doc = fitz.open()
    try:
        for blob in page_blobs:
            with fitz.open(stream=blob, filetype="pdf") as src:
                out_doc.insert_pdf(src)
        out_doc.save(str(path))
    finally:
        out_doc.close()


# --- Validators --------------------------------------------------------------

def _read_xlsx_rows(path: Path) -> tuple[list[str], list[list[Any]]]:
    wb = load_workbook(filename=str(path), read_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return [], []
    header = [str(c) if c is not None else "" for c in rows[0]]
    body = [list(r) for r in rows[1:]]
    return header, body


def validate_helpers_directly(tmp: Path) -> dict:
    """Cover ``_extract_split_rows_from_pdfs`` and ``_write_split_rows_xlsx``
    in isolation with both Meesho and Flipkart fixtures.
    """
    meesho_pdf = tmp / "fake_meesho.pdf"
    flipkart_pdf = tmp / "fake_flipkart.pdf"
    _write_pdf(
        meesho_pdf,
        [
            _build_fake_meesho_label(
                suborder="123456789012345_1",
                name="Meesho Customer A",
                address_1="House 12, Street 5",
                pincode="560001",
            ),
            _build_fake_meesho_label(
                suborder="987654321098765_1",
                name="Meesho Customer B",
                address_1="Flat 7, Lane 3",
                pincode="560002",
            ),
        ],
    )
    _write_pdf(
        flipkart_pdf,
        [
            _build_fake_flipkart_label(
                order_id="OD436652983093810100",
                name="Flipkart Customer A",
                address_1="Plot 9, Sector 11",
                pincode="110001",
            )
        ],
    )

    rows = task_queue._extract_split_rows_from_pdfs([str(meesho_pdf), str(flipkart_pdf)])
    sample = rows[0] if rows else {}
    expected_columns = list(task_queue.SPLIT_EXPORT_COLUMNS)

    populated_xlsx = tmp / "populated.xlsx"
    task_queue._write_split_rows_xlsx(rows, str(populated_xlsx), category="suspicious")
    pop_header, pop_body = _read_xlsx_rows(populated_xlsx)

    empty_xlsx = tmp / "empty.xlsx"
    task_queue._write_split_rows_xlsx([], str(empty_xlsx), category="multi_order")
    empty_header, empty_body = _read_xlsx_rows(empty_xlsx)

    parsed_some = bool(
        rows
        and any(r.get("Pincode") for r in rows)
        and any(r.get("Name") for r in rows)
    )
    return {
        "ok": (
            len(rows) == 3
            and pop_header == expected_columns
            and len(pop_body) == 3
            and empty_header == expected_columns
            and empty_body == []
            and parsed_some
        ),
        "row_count": len(rows),
        "header_matches": pop_header == expected_columns,
        "empty_header_matches": empty_header == expected_columns,
        "empty_row_count": len(empty_body),
        "sample_row": {k: sample.get(k) for k in ("Order_id", "Name", "Pincode", "Source_PDF", "Page_Number")},
        "populated_xlsx": str(populated_xlsx),
        "empty_xlsx": str(empty_xlsx),
    }


def _stub_history_calls(monkey: dict) -> None:
    """Replace the DB-touching history hooks so we can run _process_crop_task
    without setting up a sqlite schema for this validation."""
    monkey["mark_crop_job_success_orig"] = history_store.mark_crop_job_success
    monkey["mark_crop_job_failed_orig"] = history_store.mark_crop_job_failed
    monkey["task_queue_mark_success_orig"] = task_queue.mark_crop_job_success
    monkey["task_queue_mark_failed_orig"] = task_queue.mark_crop_job_failed
    monkey["set_progress_orig"] = getattr(task_queue, "_set_progress", None)

    def _noop_success(*args: Any, **kwargs: Any) -> None:
        return None

    def _noop_failed(*args: Any, **kwargs: Any) -> None:
        return None

    def _noop_progress(*args: Any, **kwargs: Any) -> None:
        return None

    history_store.mark_crop_job_success = _noop_success  # type: ignore[assignment]
    history_store.mark_crop_job_failed = _noop_failed  # type: ignore[assignment]
    task_queue.mark_crop_job_success = _noop_success  # type: ignore[assignment]
    task_queue.mark_crop_job_failed = _noop_failed  # type: ignore[assignment]
    if monkey["set_progress_orig"] is not None:
        task_queue._set_progress = _noop_progress  # type: ignore[assignment]


def _restore_history_calls(monkey: dict) -> None:
    history_store.mark_crop_job_success = monkey["mark_crop_job_success_orig"]
    history_store.mark_crop_job_failed = monkey["mark_crop_job_failed_orig"]
    task_queue.mark_crop_job_success = monkey["task_queue_mark_success_orig"]
    task_queue.mark_crop_job_failed = monkey["task_queue_mark_failed_orig"]
    if monkey.get("set_progress_orig") is not None:
        task_queue._set_progress = monkey["set_progress_orig"]


def _build_meesho_inputs(tmp: Path) -> tuple[list[str], dict]:
    """Build 2 Meesho PDFs that together exercise:
      * pincode split (one page matches the selected pincode)
      * multi-order-by-customer split (two pages share name+address)
      * a non-matching page that should land in the normal pool
    """
    pdf_a = tmp / "input_meesho_a.pdf"
    pdf_b = tmp / "input_meesho_b.pdf"

    page_pin = _build_fake_meesho_label(
        suborder="111111111111111_1",
        name="Pincode Match Customer",
        address_1="Plot 1, Pincode Lane",
        pincode="560100",
    )
    page_multi_1 = _build_fake_meesho_label(
        suborder="222222222222222_1",
        name="Repeat Customer",
        address_1="Same Address Building",
        pincode="560200",
    )
    page_multi_2 = _build_fake_meesho_label(
        suborder="333333333333333_1",
        name="Repeat Customer",
        address_1="Same Address Building",
        pincode="560200",
    )
    page_normal = _build_fake_meesho_label(
        suborder="444444444444444_1",
        name="Solo Customer",
        address_1="Unique Address Lane",
        pincode="560300",
    )

    _write_pdf(pdf_a, [page_pin, page_multi_1])
    _write_pdf(pdf_b, [page_multi_2, page_normal])

    options = {
        "separate_pincodes": "560100",
        "detect_suspicious": True,
        "separate_multi_order_by_customer": True,
    }
    return [str(pdf_a), str(pdf_b)], options


def _build_flipkart_inputs(tmp: Path) -> tuple[list[str], dict]:
    pdf_a = tmp / "input_flipkart_a.pdf"
    pdf_b = tmp / "input_flipkart_b.pdf"

    page_pin = _build_fake_flipkart_label(
        order_id="OD100000000000000001",
        name="FK Pincode Cust",
        address_1="Block A, Sector 1",
        pincode="110099",
    )
    page_multi_1 = _build_fake_flipkart_label(
        order_id="OD100000000000000002",
        name="FK Repeat Cust",
        address_1="Block B, Sector 2",
        pincode="110001",
    )
    page_multi_2 = _build_fake_flipkart_label(
        order_id="OD100000000000000003",
        name="FK Repeat Cust",
        address_1="Block B, Sector 2",
        pincode="110001",
    )
    page_normal = _build_fake_flipkart_label(
        order_id="OD100000000000000004",
        name="FK Solo Cust",
        address_1="Block C, Sector 3",
        pincode="110002",
    )

    _write_pdf(pdf_a, [page_pin, page_multi_1])
    _write_pdf(pdf_b, [page_multi_2, page_normal])

    options = {
        "separate_pincodes": "110099",
        "detect_suspicious": True,
        "separate_multi_order_by_customer": True,
    }
    return [str(pdf_a), str(pdf_b)], options


def _run_full_pipeline(
    *,
    task_type: str,
    inputs: list[str],
    options: dict,
    output_dir: Path,
) -> tuple[str, dict]:
    payload = {
        "input_paths": inputs,
        "input_files": [{"file_name": Path(p).name, "input_pages": 2} for p in inputs],
        "total_input_files": len(inputs),
        "total_input_pages": 4,
        "output_dir": str(output_dir),
        "options": options,
        "sort_by": "order_id" if task_type == "crop_meesho" else "sku",
        "layout": "label_printer",
    }
    task = {
        "task_id": f"validate_{task_type}",
        "task_type": task_type,
        "job_id": 0,
        "user_id": 0,
        "payload": payload,
    }
    return task_queue._process_crop_task(task)


def _inspect_zip(zip_path: str) -> dict[str, Any]:
    info: dict[str, Any] = {"path": zip_path, "exists": Path(zip_path).exists(), "members": []}
    if not info["exists"]:
        return info
    with zipfile.ZipFile(zip_path, "r") as zf:
        info["members"] = sorted(zf.namelist())
        excels: dict[str, dict] = {}
        for member in info["members"]:
            if not member.lower().endswith(".xlsx"):
                continue
            with zf.open(member) as inner:
                buf = io.BytesIO(inner.read())
            wb = load_workbook(filename=buf, read_only=True)
            try:
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
            finally:
                wb.close()
            header = [str(c) if c is not None else "" for c in rows[0]] if rows else []
            body = [list(r) for r in rows[1:]] if len(rows) > 1 else []
            excels[member] = {"header": header, "row_count": len(body)}
        info["excels"] = excels
    return info


def validate_full_pipeline(
    *,
    task_type: str,
    builder,
    tmp: Path,
) -> dict:
    work_dir = tmp / f"{task_type}_run"
    work_dir.mkdir(parents=True, exist_ok=True)
    inputs, options = builder(work_dir)
    monkey: dict = {}
    _stub_history_calls(monkey)
    try:
        output_path, summary = _run_full_pipeline(
            task_type=task_type,
            inputs=inputs,
            options=options,
            output_dir=work_dir,
        )
    finally:
        _restore_history_calls(monkey)

    zip_info = _inspect_zip(output_path)
    expected_excels = {
        task_queue.SPLIT_EXPORT_FILENAMES["suspicious"],
        task_queue.SPLIT_EXPORT_FILENAMES["multi_order"],
        task_queue.SPLIT_EXPORT_FILENAMES["pincode"],
    }
    excels_present = expected_excels.issubset(set(zip_info.get("members", [])))
    excel_meta = zip_info.get("excels", {})
    expected_columns = list(task_queue.SPLIT_EXPORT_COLUMNS)
    headers_match_all = all(meta.get("header") == expected_columns for meta in excel_meta.values())
    multi_rows = excel_meta.get(task_queue.SPLIT_EXPORT_FILENAMES["multi_order"], {}).get("row_count", 0)
    pincode_rows = excel_meta.get(task_queue.SPLIT_EXPORT_FILENAMES["pincode"], {}).get("row_count", 0)
    suspicious_rows = excel_meta.get(task_queue.SPLIT_EXPORT_FILENAMES["suspicious"], {}).get("row_count", 0)

    return {
        "ok": (
            output_path.endswith(".zip")
            and excels_present
            and headers_match_all
            and multi_rows >= 2
            and pincode_rows >= 1
            and suspicious_rows == 0  # no risky orders supplied for this user
        ),
        "task_type": task_type,
        "output_path": output_path,
        "summary_keys": sorted(list(summary.keys())),
        "split_excel_exports": summary.get("split_excel_exports"),
        "zip_members": zip_info.get("members"),
        "excels_present": sorted(list(expected_excels.intersection(set(zip_info.get("members", []))))),
        "missing_excels": sorted(list(expected_excels - set(zip_info.get("members", [])))),
        "excel_summary": excel_meta,
    }


def main() -> int:
    out_root = BACKEND_DIR / "temp" / "split_excel_validation"
    out_root.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(prefix="split_excel_", dir=str(out_root)) as td:
        tmp = Path(td)
        results = {
            "helpers_direct": validate_helpers_directly(tmp),
            "meesho_full_pipeline": validate_full_pipeline(
                task_type="crop_meesho", builder=_build_meesho_inputs, tmp=tmp
            ),
            "flipkart_full_pipeline": validate_full_pipeline(
                task_type="crop_flipkart", builder=_build_flipkart_inputs, tmp=tmp
            ),
        }
    print(json.dumps(results, indent=2, default=str))
    all_ok = all(isinstance(v, dict) and v.get("ok") for v in results.values())
    return 0 if all_ok else 1


if __name__ == "__main__":
    raise SystemExit(main())

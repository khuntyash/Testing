from __future__ import annotations

import io
import csv
import os
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import fitz
import pytesseract
from openpyxl import Workbook
from PIL import Image

HEADERS = [
    "Order_id",
    "Name",
    "Address_1",
    "Address_2",
    "Address_3",
    "District",
    "State",
    "Pincode",
    "Sku",
    "Size",
    "Quantity",
    "Payment_Mode",
    "Courier_Partner",
    "Courier_trans_id",
    "Processed_At",
]

COLUMN_PRESETS: dict[str, list[tuple[str, str | None]]] = {
    "standard_v1": [(h, h) for h in HEADERS],
    "reference_v2": [
        ("Order Id", "Order_id"),
        ("Customer Name", "Name"),
        ("Address Line 1", "Address_1"),
        ("Address Line 2", "Address_2"),
        ("Address Line 3", "Address_3"),
        ("District", "District"),
        ("State", "State"),
        ("Pincode", "Pincode"),
        ("SKU", "Sku"),
        ("Size", "Size"),
        ("Quantity", "Quantity"),
        ("Payment Mode", "Payment_Mode"),
        ("Courier Partner", "Courier_Partner"),
        ("Courier Trans Id", "Courier_trans_id"),
        ("Processed At", "Processed_At"),
    ],
    "compact_v1": [
        ("Order_id", "Order_id"),
        ("Sku", "Sku"),
        ("Size", "Size"),
        ("Quantity", "Quantity"),
        ("Name", "Name"),
        ("Pincode", "Pincode"),
        ("Payment_Mode", "Payment_Mode"),
        ("Courier_Partner", "Courier_Partner"),
        ("Courier_trans_id", "Courier_trans_id"),
        ("Processed_At", "Processed_At"),
    ],
}

REQUIRED_FIELDS = ("Order_id", "Name", "Pincode")


# ---------------------------------------------------------------------------
# Flipkart label support
# ---------------------------------------------------------------------------

FLIPKART_DEFAULT_COURIER_PARTNER = "E-Kart Logistics"

# Indian state/UT codes used in the trailing "IN-XX" suffix on Flipkart labels.
FLIPKART_STATE_CODES: dict[str, str] = {
    "AP": "Andhra Pradesh",
    "AR": "Arunachal Pradesh",
    "AS": "Assam",
    "BR": "Bihar",
    "CT": "Chhattisgarh",
    "CG": "Chhattisgarh",
    "GA": "Goa",
    "GJ": "Gujarat",
    "HR": "Haryana",
    "HP": "Himachal Pradesh",
    "JH": "Jharkhand",
    "KA": "Karnataka",
    "KL": "Kerala",
    "MP": "Madhya Pradesh",
    "MH": "Maharashtra",
    "MN": "Manipur",
    "ML": "Meghalaya",
    "MZ": "Mizoram",
    "NL": "Nagaland",
    "OR": "Odisha",
    "OD": "Odisha",
    "PB": "Punjab",
    "RJ": "Rajasthan",
    "SK": "Sikkim",
    "TN": "Tamil Nadu",
    "TG": "Telangana",
    "TS": "Telangana",
    "TR": "Tripura",
    "UP": "Uttar Pradesh",
    "UT": "Uttarakhand",
    "UK": "Uttarakhand",
    "WB": "West Bengal",
    "AN": "Andaman and Nicobar Islands",
    "CH": "Chandigarh",
    "DN": "Dadra and Nagar Haveli and Daman and Diu",
    "DD": "Daman and Diu",
    "DL": "Delhi",
    "JK": "Jammu and Kashmir",
    "LA": "Ladakh",
    "LD": "Lakshadweep",
    "PY": "Puducherry",
}


class OcrSetupError(RuntimeError):
    """Raised when OCR runtime is unavailable on server machine."""


def _configure_tesseract_binary() -> None:
    # Windows winget installs here by default; set explicitly when PATH is missing.
    default_windows_path = Path(r"C:\Program Files\Tesseract-OCR\tesseract.exe")
    if os.name == "nt" and default_windows_path.exists():
        pytesseract.pytesseract.tesseract_cmd = str(default_windows_path)


_configure_tesseract_binary()


def _clean_line(line: str) -> str:
    return re.sub(r"\s+", " ", line).strip().rstrip(",")


def _safe_float(value: str | int | float | None) -> float:
    try:
        return float(value)
    except Exception:
        return -1.0


def _normalize_size(value: str) -> str:
    cleaned = re.sub(r"\s+", " ", str(value or "").strip())
    if not cleaned:
        return ""
    upper = cleaned.upper()
    if upper in {"FREE SIZE", "FREESIZE"}:
        return "Free Size"
    if upper in {"ONE SIZE", "ONESIZE"}:
        return "One Size"
    if re.fullmatch(r"\d{1,2}\s*[-/]\s*\d{1,2}\s*(YEARS?|YRS?|MONTHS?|MOS?)", upper):
        normalized = re.sub(r"\s+", " ", upper)
        normalized = normalized.replace("YRS", "YEARS").replace("MOS", "MONTHS")
        return normalized.title()
    return upper


def _normalize_quantity(value: str) -> str:
    if value is None:
        return ""
    match = re.search(r"\d{1,3}", str(value))
    if not match:
        return ""
    qty = int(match.group(0))
    if qty <= 0:
        return ""
    return str(qty)


def _extract_size_and_quantity(text: str, lines: list[str]) -> tuple[str, str]:
    size = ""
    quantity = ""

    for line in lines:
        if "size" not in line.lower():
            continue
        # Stop capture before qty/quantity or separators on the same line.
        match = re.search(
            r"\bsize\s*[:\-]?\s*([A-Za-z0-9][A-Za-z0-9\s\-\/]{0,16}?)(?=\s+\b(?:qty|quantity)\b|[,|]|$)",
            line,
            flags=re.IGNORECASE,
        )
        if not match:
            continue
        candidate = _normalize_size(match.group(1))
        if candidate:
            size = candidate
            break
    if not size:
        for pattern in [
            r"\b(free\s*size|one\s*size)\b",
            r"\b(\d{1,2}\s*[-/]\s*\d{1,2}\s*(?:years?|yrs?|months?|mos?))\b",
            r"\b(XXXS|XXS|XS|S|M|L|XL|XXL|XXXL|4XL|5XL)\b",
        ]:
            match = re.search(pattern, text, flags=re.IGNORECASE)
            if not match:
                continue
            candidate = _normalize_size(match.group(1))
            if candidate:
                size = candidate
                break
    # Table fallback commonly seen in Meesho labels:
    # Product Details | SKU | Size | Qty | Color | Order No.
    if not size:
        for idx, line in enumerate(lines):
            if _normalized_key(line) != "size":
                continue
            window_before = " ".join(lines[max(0, idx - 6) : idx + 1]).lower()
            if not any(token in window_before for token in ("product details", "sku", "qty", "order no")):
                continue
            for candidate_line in lines[idx + 1 : idx + 12]:
                candidate_line = candidate_line.strip()
                if not candidate_line:
                    continue
                if re.fullmatch(r"[A-Za-z]{1,5}", candidate_line):
                    size = _normalize_size(candidate_line)
                    break
                if re.fullmatch(
                    r"\d{1,2}\s*[-/]\s*\d{1,2}\s*(?:years?|yrs?|months?|mos?)",
                    candidate_line,
                    flags=re.IGNORECASE,
                ):
                    size = _normalize_size(candidate_line)
                    break
            if size:
                break

    qty_patterns = [
        r"\b(\d{1,3})\s*[xX]\s*(?:rs|inr|mrp)?\.?\s*[\d,]+(?:\.\d+)?\b",
        r"\b(?:qty|quantity)\s*[:\-]?\s*(\d{1,3})\b",
        r"\b(\d{1,3})\s*(?:pcs?|pieces?|items?)\b",
    ]
    for pattern in qty_patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if not match:
            continue
        candidate = _normalize_quantity(match.group(1))
        if candidate:
            quantity = candidate
            break

    # If quantity wasn't found, try extracting from product-details style rows.
    if not quantity:
        for line in lines:
            if "qty" not in line.lower() and "quantity" not in line.lower():
                continue
            match = re.search(r"\b(\d{1,3})\b", line)
            if match:
                quantity = _normalize_quantity(match.group(1))
                if quantity:
                    break

    # Table fallback commonly seen in Meesho labels:
    # Product Details | SKU | Size | Qty | Color | Order No.
    if not quantity:
        for idx, line in enumerate(lines):
            if _normalized_key(line) != "qty":
                continue
            window_before = " ".join(lines[max(0, idx - 6) : idx + 1]).lower()
            if not any(token in window_before for token in ("product details", "sku", "size", "order no")):
                continue
            for candidate_line in lines[idx + 1 : idx + 12]:
                candidate_line = candidate_line.strip()
                if not candidate_line:
                    continue
                if re.fullmatch(r"\d{1,2}", candidate_line):
                    quantity = _normalize_quantity(candidate_line)
                    if quantity:
                        break
            if quantity:
                break

    return size, quantity


def _extract_page_text(page: fitz.Page) -> tuple[str, str, float]:
    text = page.get_text("text").strip()
    if text:
        return text, "native", 100.0
    pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
    mode = "RGB" if pix.alpha == 0 else "RGBA"
    image = Image.frombytes(mode, [pix.width, pix.height], pix.samples)
    if mode == "RGBA":
        image = image.convert("RGB")
    try:
        ocr_text = pytesseract.image_to_string(image).strip()
        data = pytesseract.image_to_data(image, output_type=pytesseract.Output.DICT)
        conf_values = [_safe_float(v) for v in data.get("conf", [])]
        conf_values = [v for v in conf_values if v >= 0]
        confidence = round(sum(conf_values) / len(conf_values), 2) if conf_values else 0.0
        return ocr_text, "ocr", confidence
    except pytesseract.TesseractNotFoundError as exc:
        raise OcrSetupError(
            "Tesseract OCR is not installed on the server. Install Tesseract and restart backend."
        ) from exc


def _is_flipkart_label(text: str) -> bool:
    """Detect Flipkart shipping labels using multiple positive signals.

    Score-based check is used to stay robust against partial OCR noise; the
    parser falls back to Meesho behaviour when the score is low.
    """
    if not text:
        return False
    score = 0
    if re.search(r"E[\s\-]*Kart\s+Logistics", text, flags=re.IGNORECASE):
        score += 2
    if re.search(r"\bOD\d{14,22}\b", text):
        score += 2
    if re.search(r"SKU\s*ID\s*\|\s*Description", text, flags=re.IGNORECASE):
        score += 1
    if re.search(r"Shipping\s*/\s*Customer\s*address", text, flags=re.IGNORECASE):
        score += 1
    if re.search(r"\bAWB\s*No\.", text, flags=re.IGNORECASE):
        score += 1
    if re.search(r"\bIN-[A-Z]{2}\b", text):
        score += 1
    return score >= 3


def _flipkart_extract_order_and_payment(text: str) -> tuple[str, str]:
    """Return (order_id, payment_mode) for a Flipkart label.

    Order ID lines look like: ``OD436930471027953400  COD`` and may show
    PREPAID, COD, EXCHANGE etc. as the suffix.
    """
    order_id = ""
    payment_mode = ""
    pair = re.search(
        r"\b(OD\d{14,22})\b\s*(COD|PREPAID|PARTIAL\s*COD|EXCHANGE|REPLACE(?:MENT)?)\b",
        text,
        flags=re.IGNORECASE,
    )
    if pair:
        order_id = pair.group(1).strip()
        token = re.sub(r"\s+", "", pair.group(2)).upper()
        if token == "COD":
            payment_mode = "COD"
        elif token == "PREPAID":
            payment_mode = "Prepaid"
        elif token == "PARTIALCOD":
            payment_mode = "Partial COD"
        elif token in {"EXCHANGE", "REPLACE", "REPLACEMENT"}:
            payment_mode = "Exchange"
    if not order_id:
        only = re.search(r"\b(OD\d{14,22})\b", text)
        if only:
            order_id = only.group(1).strip()
    if not payment_mode:
        text_lower = text.lower()
        if re.search(r"\bprepaid\b", text_lower):
            payment_mode = "Prepaid"
        elif re.search(r"\bcod\b", text_lower):
            payment_mode = "COD"
    return order_id, payment_mode


def _flipkart_extract_awb(text: str) -> str:
    """Extract the AWB / Courier_trans_id token from a Flipkart label."""
    m = re.search(r"AWB\s*No\.\s*([A-Z0-9]{6,})", text, flags=re.IGNORECASE)
    if m:
        return m.group(1).strip()
    fallback = re.search(r"\b(FMP[CP]?\d{8,14}|SF[A-Z0-9]{8,16})\b", text)
    if fallback:
        return fallback.group(1).strip()
    return ""


_AWB_TOKEN_RE = re.compile(r"^(?:FMP[CP]?\d{8,14}|SF[A-Z0-9]{8,16})$")


def _flipkart_sku_block(lines: list[str]) -> list[str]:
    """Return the lines that hold SKU rows on a Flipkart label.

    The block is bounded above by the standalone ``QTY`` header line and
    below by the AWB tracking token / ``STD`` / ``AWB No.`` line.
    """
    qty_idx = -1
    for i, line in enumerate(lines):
        if line.strip().upper() == "QTY":
            qty_idx = i
            break
    if qty_idx == -1:
        return []
    end_idx = len(lines)
    for j in range(qty_idx + 1, len(lines)):
        s = lines[j].strip()
        if _AWB_TOKEN_RE.match(s):
            end_idx = j
            break
        upper = s.upper()
        if upper == "STD" or upper.startswith("AWB NO"):
            end_idx = j
            break
        if upper.startswith("ORDERED THROUGH"):
            end_idx = j
            break
    return lines[qty_idx + 1:end_idx]


def _flipkart_parse_sku_rows(block: list[str]) -> tuple[list[str], int]:
    """Parse SKU rows. Returns (skus, total_quantity).

    Rows usually fit one line (``1 SKU | description``) but long SKU codes
    occasionally wrap before the ``|`` divider. We merge such continuations
    so the SKU value remains intact.
    """
    skus: list[str] = []
    total_qty = 0
    if not block:
        return skus, total_qty

    row_start_re = re.compile(r"^\s*\d+\s+\S")
    capture_re = re.compile(r"^\s*\d+\s+(.+?)\s*\|\s*(.*)$")

    row_starts = [i for i, line in enumerate(block) if row_start_re.match(line)]
    if not row_starts:
        return skus, total_qty
    row_starts.append(len(block))

    for r_idx in range(len(row_starts) - 1):
        start = row_starts[r_idx]
        end = row_starts[r_idx + 1]
        first_line = block[start]
        consumed = 1
        # If `|` is missing on the head line, glue subsequent lines until found.
        if "|" not in first_line:
            for k in range(start + 1, end):
                first_line = first_line.rstrip() + " " + block[k].strip()
                consumed += 1
                if "|" in first_line:
                    break
        m = capture_re.match(first_line)
        if not m:
            continue
        sku = m.group(1).strip().rstrip(",")
        skus.append(sku)

        row_qty = 0
        same_line_tail = m.group(2).rstrip()
        same_match = re.search(r"(?:^|\s)(\d{1,3})$", same_line_tail)

        # Standalone digit on a following line within this row's slice.
        scan_start = start + consumed
        for k in range(scan_start, end):
            inner = block[k].strip()
            if re.fullmatch(r"\d{1,3}", inner):
                row_qty = int(inner)
                break
        if not row_qty and same_match:
            row_qty = int(same_match.group(1))
        if not row_qty:
            row_qty = 1
        total_qty += row_qty

    return skus, total_qty


def _flipkart_parse_address(lines: list[str]) -> dict:
    """Parse customer address block on a Flipkart label."""
    out = {
        "name": "",
        "address_1": "",
        "address_2": "",
        "address_3": "",
        "district": "",
        "state": "",
        "pincode": "",
    }
    addr_start = -1
    for i, line in enumerate(lines):
        if re.search(r"shipping\s*/\s*customer\s*address", line, flags=re.IGNORECASE):
            addr_start = i + 1
            break
    if addr_start == -1:
        return out
    addr_end = len(lines)
    for j in range(addr_start, len(lines)):
        ls = lines[j].strip().lower()
        if ls.startswith("not for resale") or ls.startswith("printed at"):
            addr_end = j
            break
    block = lines[addr_start:addr_end]
    if not block:
        return out

    # Customer name — typically "Name: <name>," on the first line. Some
    # records spill onto a 2nd line that ends with comma; cap at 2 lines.
    body_start = 1
    first = block[0]
    name_m = re.match(r"^\s*Name\s*:\s*(.*)$", first, flags=re.IGNORECASE)
    if name_m:
        first_name_text = name_m.group(1).strip()
    else:
        first_name_text = first.strip()
    if first_name_text.endswith(","):
        out["name"] = first_name_text.rstrip(",").strip()
    else:
        # Continue onto next line if it looks like a short name fragment that
        # ends with a comma. Avoid capturing actual address tokens.
        if (
            len(block) > 1
            and block[1].strip().endswith(",")
            and len(block[1].strip()) <= 40
            and not re.search(r"\d", block[1])
        ):
            out["name"] = (
                first_name_text + " " + block[1].strip().rstrip(",")
            ).strip()
            body_start = 2
        else:
            out["name"] = first_name_text.rstrip(",").strip()

    body = block[body_start:]
    if not body:
        return out

    # Locate the line(s) that carry "<District> - <pincode>, IN-XX". Some
    # garbled variants drop the dash/comma and read "<pincode> IN XX". Also
    # handle the case where "IN-" sits at the end of a line and the state
    # code spills onto the next line ("... IN-\nTN").
    pincode_re = re.compile(r"\b\d{6}[\s,]*IN[-\s]*[A-Z]{2}\b")
    pincode_line_idx = -1
    for i, b in enumerate(body):
        if pincode_re.search(b):
            pincode_line_idx = i
            break
    if pincode_line_idx == -1:
        for i in range(len(body) - 1):
            joined = body[i].rstrip() + " " + body[i + 1].strip()
            if pincode_re.search(joined):
                body = body[:i] + [joined] + body[i + 2:]
                pincode_line_idx = i
                break

    district_text = ""
    if pincode_line_idx >= 0:
        pincode_text = body[pincode_line_idx].strip()
        single = re.match(
            r"^(.+?)\s*-\s*(\d{6})\s*[,\s]\s*IN[-\s]*([A-Z]{2})\b",
            pincode_text,
        )
        if single:
            district_text = single.group(1).strip().rstrip(",")
            out["pincode"] = single.group(2)
            code = single.group(3).upper()
            out["state"] = FLIPKART_STATE_CODES.get(code, code)
        else:
            split_m = re.search(
                r"(\d{6})\s*[,\s]\s*IN[-\s]*([A-Z]{2})\b", pincode_text
            )
            if split_m:
                out["pincode"] = split_m.group(1)
                code = split_m.group(2).upper()
                out["state"] = FLIPKART_STATE_CODES.get(code, code)
            # Pincode line is split: prior line ends with "-".
            if pincode_line_idx > 0:
                prior = body[pincode_line_idx - 1].strip()
                if prior.endswith("-"):
                    district_text = prior[:-1].strip().rstrip(",")
                    pincode_line_idx -= 1
            # Or district may be inline before the digits.
            if not district_text:
                inline = re.match(
                    r"^(.+?)\s*[-]?\s*\d{6}\s*[,\s]\s*IN", pincode_text
                )
                if inline:
                    district_text = inline.group(1).strip().rstrip(",")

    if district_text:
        cleaned = re.sub(
            r"\b(?:District|Subdistrict|Tehsil|Taluka|Mandal|Dist\.?)\b",
            "",
            district_text,
            flags=re.IGNORECASE,
        )
        cleaned = re.sub(r"\s+", " ", cleaned).strip(" ,-")
        out["district"] = cleaned or district_text

    pure_addr_raw = body[:pincode_line_idx] if pincode_line_idx >= 0 else body
    pure_addr = []
    for line in pure_addr_raw:
        s = line.strip().rstrip(",").strip()
        if not s:
            continue
        if s.lower() in {"india", "india,"}:
            continue
        pure_addr.append(s)

    if pure_addr:
        out["address_1"] = pure_addr[0]
    if len(pure_addr) >= 2:
        out["address_2"] = pure_addr[1]
    if len(pure_addr) >= 3:
        out["address_3"] = ", ".join(pure_addr[2:])
    return out


def _parse_flipkart_fields(text: str, lines: list[str]) -> dict:
    """Extract structured fields from a Flipkart shipping label."""
    order_id, payment_mode = _flipkart_extract_order_and_payment(text)
    awb = _flipkart_extract_awb(text)
    skus, total_qty = _flipkart_parse_sku_rows(_flipkart_sku_block(lines))
    addr = _flipkart_parse_address(lines)

    sku_value = skus[0] if skus else ""
    quantity_value = str(total_qty) if total_qty else ""

    return {
        "Order_id": order_id,
        "Name": addr["name"],
        "Address_1": addr["address_1"],
        "Address_2": addr["address_2"],
        "Address_3": addr["address_3"],
        "District": addr["district"],
        "State": addr["state"],
        "Pincode": addr["pincode"],
        "Sku": sku_value,
        # Flipkart labels do not expose a dedicated size column; size is
        # encoded inside the SKU/description so we leave it blank to avoid
        # false matches.
        "Size": "",
        "Quantity": quantity_value,
        "Payment_Mode": payment_mode,
        "Courier_Partner": FLIPKART_DEFAULT_COURIER_PARTNER,
        "Courier_trans_id": awb,
        "Sold_By": _extract_sold_by(text, lines),
    }


def parse_required_fields(text: str) -> dict:
    lines = [_clean_line(line) for line in text.splitlines()]
    lines = [line for line in lines if line]

    if _is_flipkart_label(text):
        return _parse_flipkart_fields(text, lines)

    order_id = _extract_suborder_id(text)

    sku = ""
    courier_trans_id = ""
    size = ""
    quantity = ""

    payment_mode = ""
    text_lower = text.lower()
    if "prepaid: do not collect cash" in text_lower:
        payment_mode = "Prepaid"
    elif re.search(r"\bcod\b", text_lower):
        payment_mode = "COD"
    elif re.search(r"^\s*exchange\s*$", text, flags=re.IGNORECASE | re.MULTILINE):
        payment_mode = "Exchange"

    courier_partner = ""
    if "shadowfax" in text_lower:
        courier_partner = "Shadowfax"
    elif "delhivery" in text_lower:
        courier_partner = "Delhivery"
    elif re.search(r"\bxpress\s*bee?s?\b", text_lower) or "xpressbees" in text_lower:
        courier_partner = "Xpress Bees"
    elif "valmoplus" in text_lower:
        courier_partner = "ValmoPlus"
    elif re.search(r"\bvalmo\b", text_lower):
        courier_partner = "Valmo"

    name = ""
    address_1 = ""
    address_2 = ""
    address_3 = ""
    district = ""
    state = ""
    pincode = ""

    customer_start = -1
    customer_end = -1
    for idx, line in enumerate(lines):
        if line.lower() == "customer address":
            customer_start = idx + 1
            break
    if customer_start != -1:
        for idx in range(customer_start, len(lines)):
            if lines[idx].lower().startswith("if undelivered"):
                customer_end = idx
                break
        if customer_end == -1:
            customer_end = len(lines)

    address_block = lines[customer_start:customer_end] if customer_start != -1 else []
    if address_block:
        name = address_block[0]
        address_lines = address_block[1:]

        city_idx = -1
        for i, line in enumerate(address_lines):
            if re.search(r",\s*[^,]+,\s*\d{6}$", line):
                city_idx = i
                city_match = re.search(r"^([^,]+),\s*([^,]+),\s*(\d{6})$", line)
                if city_match:
                    district = city_match.group(1).strip()
                    state = city_match.group(2).strip()
                    pincode = city_match.group(3).strip()
                break

        if not pincode:
            pin_match = re.search(r"\b(\d{6})\b", " ".join(address_lines))
            if pin_match:
                pincode = pin_match.group(1)

        if not state:
            place_supply_match = re.search(r"Place of Supply:\s*([A-Za-z &]+)", text)
            if place_supply_match:
                state = place_supply_match.group(1).strip()

        pure_address_lines = address_lines[:city_idx] if city_idx != -1 else address_lines
        if pure_address_lines:
            address_1 = pure_address_lines[0]
        if len(pure_address_lines) > 1:
            address_2 = pure_address_lines[1]
        if len(pure_address_lines) > 2:
            address_3 = pure_address_lines[2]

        if not district:
            source_line = address_3 or address_2
            if source_line:
                district = source_line.split(",")[-1].strip(" -")

    if not district and address_3:
        district = address_3.split(",")[0].strip()

    if re.search(r"Nagpur District\s*,", text) and address_3.endswith("Nagpur District"):
        address_3 = f"{address_3},"

    for idx, line in enumerate(lines):
        if line.lower().startswith("order no."):
            if idx + 1 < len(lines):
                candidate = lines[idx + 1]
                if candidate and not re.fullmatch(r"\d{18}_\d+", candidate):
                    sku = candidate
            break

    if not sku and order_id:
        for idx, line in enumerate(lines):
            if order_id in line and idx >= 1:
                candidate = lines[idx - 1]
                if candidate and re.search(r"[A-Za-z]", candidate):
                    sku = candidate
                break

    size, quantity = _extract_size_and_quantity(text, lines)

    if (
        not payment_mode
        and "check the payable amount on the app" in text_lower
        and courier_partner in {"Valmo", "ValmoPlus"}
    ):
        payment_mode = "Check the payable amount on the app"
    elif courier_partner == "Valmo" and payment_mode == "COD":
        payment_mode = "Check the payable amount on the app"

    id_patterns = [
        r"\bSF[0-9A-Z]{8,}\b",
        r"\bVL\d{10,}\b",
        r"\b\d{12,20}\b",
    ]
    for pattern in id_patterns:
        match = re.search(pattern, text)
        if match:
            courier_trans_id = match.group(0)
            break

    if not courier_trans_id:
        for idx, line in enumerate(lines):
            if line.lower().startswith("product details") and idx >= 1:
                prev_line = lines[idx - 1]
                if re.fullmatch(r"[A-Z0-9]{10,}", prev_line):
                    courier_trans_id = prev_line
                break

    sold_by = _extract_sold_by(text, lines)

    return {
        "Order_id": order_id,
        "Name": name,
        "Address_1": address_1,
        "Address_2": address_2,
        "Address_3": address_3,
        "District": district,
        "State": state,
        "Pincode": pincode,
        "Sku": sku,
        "Size": size,
        "Quantity": quantity,
        "Payment_Mode": payment_mode,
        "Courier_Partner": courier_partner,
        "Courier_trans_id": courier_trans_id,
        "Sold_By": sold_by,
    }


def _extract_sold_by(text: str, lines: list[str]) -> str:
    """Best-effort extractor for the seller / "Sold By" string on a label.

    Tries (in order):
      1. Explicit "Sold by:" / "Sold by " labels (Flipkart and some Meesho
         tax-invoice lines).
      2. The first non-empty line after "If undelivered, return to:" which
         on Meesho labels carries the seller / sender name.
    Returns "" when nothing reasonable is found so the caller can keep an
    empty-string fallback in exports.
    """
    if not text:
        return ""

    inline_match = re.search(
        r"sold\s*by\s*[:\-]?\s*(.+)",
        text,
        flags=re.IGNORECASE,
    )
    if inline_match:
        candidate = inline_match.group(1).strip()
        candidate = candidate.splitlines()[0].strip()
        candidate = candidate.rstrip(",;.").strip()
        if candidate:
            return candidate

    for idx, line in enumerate(lines or []):
        if line.lower().startswith("if undelivered"):
            for follow in lines[idx + 1 : idx + 4]:
                cleaned = (follow or "").strip().rstrip(",;").strip()
                if not cleaned:
                    continue
                low = cleaned.lower()
                if low.startswith("order no") or low.startswith("return to"):
                    continue
                return cleaned
            break

    return ""


def _build_issue_list(record: dict) -> list[str]:
    issues: list[str] = []
    for key in REQUIRED_FIELDS:
        if not str(record.get(key, "")).strip():
            issues.append(f"missing_{key.lower()}")
    return issues


def _normalized_key(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (text or "").strip().lower())


def _extract_suborder_id(text: str) -> str:
    # Prefer explicit sub-order labels when present.
    explicit_patterns = [
        r"\bsub[\s_-]*order[\s_-]*id\b\s*[:#-]?\s*([A-Za-z0-9_-]{8,})",
        r"\bsub[\s_-]*order\b\s*[:#-]?\s*([A-Za-z0-9_-]{8,})",
    ]
    for pattern in explicit_patterns:
        match = re.search(pattern, text or "", flags=re.IGNORECASE)
        if match:
            return (match.group(1) or "").strip()

    # Legacy label format often used as sub-order id, e.g. 123456789012345678_1
    legacy = re.search(r"\b\d{14,24}[_-]\d+\b", text or "")
    if legacy:
        return legacy.group(0).strip()
    return ""


def _record_dedupe_key(record: dict) -> str:
    suborder_id = _normalized_key(str(record.get("Order_id", "")))
    if suborder_id:
        return f"suborder:{suborder_id}"
    return ""


def _merge_record_values(existing: dict, incoming: dict) -> dict:
    merged = dict(existing)
    for key, value in incoming.items():
        old = str(merged.get(key, "")).strip()
        new = str(value or "").strip()
        if key == "Processed_At" and new:
            # Keep latest processing date when merging duplicate suborders.
            merged[key] = value
            continue
        # Keep richer value when existing is empty or shorter.
        if (not old and new) or (new and len(new) > len(old)):
            merged[key] = value
    return merged


def deduplicate_records(records: list[dict]) -> tuple[list[dict], int]:
    """Removes duplicate invoice rows while preserving stable order.

    Strict key: Suborder ID only (stored in Order_id field).
    Rows without suborder id are kept as-is and not deduplicated.
    """
    seen: dict[str, dict] = {}
    order: list[str] = []
    removed = 0
    no_key_counter = 0
    for record in records:
        key = _record_dedupe_key(record)
        if not key:
            key = f"no_suborder:{no_key_counter}"
            no_key_counter += 1
            seen[key] = dict(record)
            order.append(key)
            continue
        if key in seen:
            seen[key] = _merge_record_values(seen[key], record)
            removed += 1
            continue
        seen[key] = dict(record)
        order.append(key)
    return [seen[k] for k in order], removed


def _resolve_column_plan(column_preset: str, custom_columns: str | None) -> list[tuple[str, str | None]]:
    alias_to_source = {
        "orderid": "Order_id",
        "name": "Name",
        "address1": "Address_1",
        "address2": "Address_2",
        "address3": "Address_3",
        "district": "District",
        "state": "State",
        "pincode": "Pincode",
        "sku": "Sku",
        "size": "Size",
        "quantity": "Quantity",
        "qty": "Quantity",
        "paymentmode": "Payment_Mode",
        "courierpartner": "Courier_Partner",
        "couriertransid": "Courier_trans_id",
        "processedat": "Processed_At",
        "printdate": "Processed_At",
    }
    custom = [c.strip() for c in (custom_columns or "").split(",") if c.strip()]
    if custom:
        plan: list[tuple[str, str | None]] = []
        for col in custom:
            source = alias_to_source.get(_normalized_key(col))
            plan.append((col, source))
        return plan
    return COLUMN_PRESETS.get(column_preset, COLUMN_PRESETS["standard_v1"])


def _extract_records_from_single_pdf(pdf_path: str | Path) -> tuple[list[dict], list[dict], dict]:
    records: list[dict] = []
    report_rows: list[dict] = []
    total_pages = 0
    ocr_pages = 0
    confidence_sum = 0.0
    confidence_count = 0
    pdf_name = Path(pdf_path).name
    doc = fitz.open(str(pdf_path))
    try:
        for page_idx, page in enumerate(doc, start=1):
            total_pages += 1
            text, source, confidence = _extract_page_text(page)
            if text.strip():
                parsed = parse_required_fields(text)
                issues = _build_issue_list(parsed)
                status = "ok" if not issues else "partial"
                records.append(parsed)
                report_rows.append(
                    {
                        "source_pdf": pdf_name,
                        "page_number": page_idx,
                        "text_source": source,
                        "ocr_confidence": confidence,
                        "parse_status": status,
                        "issues": ", ".join(issues),
                        "Order_id": parsed.get("Order_id", ""),
                        "Name": parsed.get("Name", ""),
                        "Pincode": parsed.get("Pincode", ""),
                        "Sku": parsed.get("Sku", ""),
                        "Size": parsed.get("Size", ""),
                        "Quantity": parsed.get("Quantity", ""),
                        "Courier_trans_id": parsed.get("Courier_trans_id", ""),
                    }
                )
            else:
                report_rows.append(
                    {
                        "source_pdf": pdf_name,
                        "page_number": page_idx,
                        "text_source": source,
                        "ocr_confidence": confidence,
                        "parse_status": "failed",
                        "issues": "empty_text_after_ocr",
                        "Order_id": "",
                        "Name": "",
                        "Pincode": "",
                        "Sku": "",
                        "Size": "",
                        "Quantity": "",
                        "Courier_trans_id": "",
                    }
                )
            if source == "ocr":
                ocr_pages += 1
                confidence_sum += confidence
                confidence_count += 1
    finally:
        doc.close()
    return records, report_rows, {
        "total_pages": total_pages,
        "ocr_pages": ocr_pages,
        "confidence_sum": confidence_sum,
        "confidence_count": confidence_count,
    }


def extract_records_from_pdfs(
    pdf_paths: list[str | Path],
    *,
    max_workers: int | None = None,
    progress_callback=None,
) -> tuple[list[dict], list[dict], dict]:
    records: list[dict] = []
    report_rows: list[dict] = []
    total_pages = 0
    ocr_pages = 0
    confidence_sum = 0.0
    confidence_count = 0
    processed_files = 0
    total_files = len(pdf_paths)
    safe_workers = int(max_workers or 0)
    if safe_workers <= 0:
        safe_workers = min(8, max(1, (os.cpu_count() or 4)))
    safe_workers = max(1, min(safe_workers, max(1, total_files)))

    with ThreadPoolExecutor(max_workers=safe_workers) as executor:
        futures = {executor.submit(_extract_records_from_single_pdf, p): p for p in pdf_paths}
        for fut in as_completed(futures):
            part_records, part_report, part_stats = fut.result()
            records.extend(part_records)
            report_rows.extend(part_report)
            total_pages += int(part_stats.get("total_pages") or 0)
            ocr_pages += int(part_stats.get("ocr_pages") or 0)
            confidence_sum += float(part_stats.get("confidence_sum") or 0)
            confidence_count += int(part_stats.get("confidence_count") or 0)
            processed_files += 1
            if callable(progress_callback):
                try:
                    progress_callback(
                        {
                            "processed_files": processed_files,
                            "total_files": total_files,
                            "total_pages": total_pages,
                            "ocr_pages": ocr_pages,
                            "records": len(records),
                        }
                    )
                except Exception:
                    pass

    report_rows.sort(key=lambda r: (str(r.get("source_pdf", "")), int(r.get("page_number", 0))))

    summary = {
        "total_files": total_files,
        "processed_files": processed_files,
        "total_pages": total_pages,
        "total_records": len(records),
        "ocr_pages": ocr_pages,
        "failed_pages": len([r for r in report_rows if r.get("parse_status") == "failed"]),
        "partial_pages": len([r for r in report_rows if r.get("parse_status") == "partial"]),
        "average_ocr_confidence": round(confidence_sum / confidence_count, 2) if confidence_count else 0.0,
    }
    return records, report_rows, summary


def build_excel_bytes(
    records: list[dict],
    report_rows: list[dict],
    summary: dict,
    *,
    column_preset: str = "standard_v1",
    custom_columns: str = "",
) -> bytes:
    column_plan = _resolve_column_plan(column_preset, custom_columns)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append([col for col, _ in column_plan])
    for row in records:
        ws.append([row.get(source, "") if source else "" for _, source in column_plan])

    report_ws = wb.create_sheet("OCR_Report")
    report_headers = [
        "source_pdf",
        "page_number",
        "text_source",
        "ocr_confidence",
        "parse_status",
        "issues",
        "Order_id",
        "Name",
        "Pincode",
        "Sku",
        "Size",
        "Quantity",
        "Courier_trans_id",
    ]
    report_ws.append(report_headers)
    for row in report_rows:
        report_ws.append([row.get(h, "") for h in report_headers])

    summary_ws = wb.create_sheet("Summary")
    summary_ws.append(["metric", "value"])
    for key in (
        "total_pages",
        "total_records",
        "ocr_pages",
        "partial_pages",
        "failed_pages",
        "average_ocr_confidence",
    ):
        summary_ws.append([key, summary.get(key, "")])
    summary_ws.append(["column_preset", column_preset or "standard_v1"])
    summary_ws.append(["custom_columns", custom_columns or ""])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def build_csv_bytes(
    records: list[dict],
    *,
    column_preset: str = "standard_v1",
    custom_columns: str = "",
) -> bytes:
    column_plan = _resolve_column_plan(column_preset, custom_columns)
    out = io.StringIO()
    writer = csv.writer(out, lineterminator="\n")
    writer.writerow([col for col, _ in column_plan])
    for row in records:
        writer.writerow([row.get(source, "") if source else "" for _, source in column_plan])
    return out.getvalue().encode("utf-8")

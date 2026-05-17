from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook


REQUIRED_RETURN_COLUMNS = [
    "Suborder Number",
    "AWB Number",
    "Type of Return",
    "Sub Type",
    "Status",
    "Return Reason",
    "Detailed Return Reason",
]

# Flipkart Returns Excel format (from Flipkart Seller Dashboard).
# The workbook ships with a "Help" sheet followed by a "Returns" data sheet.
# Header keys are lowercase snake_case and differ from Meesho's human-readable
# column names, so we normalize them into the canonical REQUIRED_RETURN_COLUMNS
# shape before the downstream analyzer consumes them.
FLIPKART_RETURN_HEADERS = {
    "return_id",
    "order_item_id",
    "fulfilment_type",
    "return_status",
    "return_reason",
    "return_sub_reason",
    "return_type",
    "return_result",
    "reverse_logistics_tracking_id",
    "sku",
    "quantity",
}
FLIPKART_REQUIRED_COLS = [
    "order_item_id",
    "return_reason",
    "return_status",
    "return_type",
]


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_suborder(value: object) -> str:
    text = _normalize_text(value).lower()
    if not text:
        return ""
    for prefix in ("oi:", "od:"):
        if text.startswith(prefix):
            text = text[len(prefix):]
            break
    return text


def _strip_flipkart_prefix(value: object, *, prefixes: tuple[str, ...]) -> str:
    text = _normalize_text(value)
    if not text:
        return text
    lower = text.lower()
    for prefix in prefixes:
        if lower.startswith(prefix):
            return text[len(prefix):].strip()
    return text


def _humanize_flipkart_reason(value: object) -> str:
    """Flipkart exposes reason codes in snake/pascal case.

    Humanize them for UI surfaces while keeping the raw tokens recognizable
    to downstream risk classifiers (which do substring matches on lowercased
    text).
    """
    text = _normalize_text(value)
    if not text:
        return text
    return text.replace("_", " ").strip()


def _map_flipkart_return_type(value: object) -> str:
    """Map Flipkart return_type values to the canonical `Type of Return` token.

    - `courier_return` -> `RTO`
    - `customer_return` -> `Customer Return`
    Anything else is passed through humanized so risk classifiers can still
    substring-match the raw token when needed.
    """
    text = _normalize_text(value).lower()
    if not text:
        return ""
    if "courier" in text:
        return "RTO"
    if "customer" in text:
        return "Customer Return"
    return _humanize_flipkart_reason(value)


def _convert_flipkart_row(raw: dict) -> dict | None:
    """Normalize one Flipkart returns row to the canonical schema."""
    suborder = _strip_flipkart_prefix(raw.get("order_item_id", ""), prefixes=("oi:",))
    if not suborder:
        return None
    awb = _strip_flipkart_prefix(
        raw.get("reverse_logistics_tracking_id", ""), prefixes=("rtr:",)
    )
    return_reason = _humanize_flipkart_reason(raw.get("return_reason", ""))
    sub_reason = _humanize_flipkart_reason(raw.get("return_sub_reason", ""))
    status = _humanize_flipkart_reason(raw.get("return_status", ""))
    type_of_return = _map_flipkart_return_type(raw.get("return_type", ""))
    detailed = _humanize_flipkart_reason(
        raw.get("detailed_pv_output", "") or raw.get("primary_pv_output", "") or sub_reason
    )
    return {
        "Suborder Number": suborder,
        "AWB Number": awb,
        "Type of Return": type_of_return,
        "Sub Type": sub_reason or _humanize_flipkart_reason(raw.get("return_result", "")),
        "Status": status,
        "Return Reason": return_reason,
        "Detailed Return Reason": detailed,
    }


def _read_orders_csv(path: str | Path) -> list[dict]:
    p = Path(path)
    if not p.exists():
        raise ValueError("Order CSV not found on server.")
    rows: list[dict] = []
    with p.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if not row:
                continue
            rows.append({str(k or "").strip(): _normalize_text(v) for k, v in row.items()})
    if not rows:
        raise ValueError("Order CSV is empty.")
    if "Order_id" not in rows[0]:
        raise ValueError("Order CSV missing required column: Order_id")
    return rows


def _pick_returns_worksheet(wb) -> object:
    """Prefer a sheet named "Returns" (Flipkart exports prepend a Help sheet)."""
    preferred = ("Returns", "returns", "Return", "return")
    for name in preferred:
        if name in wb.sheetnames:
            ws = wb[name]
            # Some Flipkart exports omit the <dimension> tag so openpyxl's
            # read-only cursor reports max_row=1. Reset forces a full scan.
            reset = getattr(ws, "reset_dimensions", None)
            if callable(reset):
                try:
                    reset()
                except Exception:
                    pass
            return ws
    ws = wb.active
    reset = getattr(ws, "reset_dimensions", None)
    if callable(reset):
        try:
            reset()
        except Exception:
            pass
    return ws


def _iter_header_and_rows(ws):
    """Yield (headers, row_values) for a worksheet, tolerating blank lead rows.

    Returns an empty (None, iterator) when no non-empty header row is found.
    """
    values_iter = ws.iter_rows(values_only=True)
    header_row = None
    for candidate in values_iter:
        if not candidate:
            continue
        if not any(c is not None and str(c).strip() for c in candidate):
            continue
        header_row = candidate
        break
    if header_row is None:
        return None, iter(())
    headers = [str(h or "").strip() for h in header_row]
    return headers, values_iter


def _is_flipkart_returns_headers(headers: list[str]) -> bool:
    if not headers:
        return False
    lowered = {h.lower() for h in headers}
    overlap = len(FLIPKART_RETURN_HEADERS & lowered)
    must_have = all(col in lowered for col in FLIPKART_REQUIRED_COLS)
    return bool(must_have and overlap >= 4)


def _read_returns_excel(path: str | Path) -> list[dict]:
    p = Path(path)
    if not p.exists():
        raise ValueError("Return Excel file not found.")
    wb = load_workbook(filename=str(p), read_only=True, data_only=True)
    try:
        ws = _pick_returns_worksheet(wb)
        headers, values_iter = _iter_header_and_rows(ws)
        if not headers:
            raise ValueError("Return Excel has no header row.")

        if _is_flipkart_returns_headers(headers):
            header_map = {h.lower(): idx for idx, h in enumerate(headers)}
            rows: list[dict] = []
            for row_vals in values_iter:
                if not row_vals or not any(
                    c is not None and str(c).strip() for c in row_vals
                ):
                    continue
                raw = {
                    key.lower(): _normalize_text(row_vals[idx] if idx < len(row_vals) else "")
                    for key, idx in header_map.items()
                }
                converted = _convert_flipkart_row(raw)
                if converted is None:
                    continue
                rows.append(converted)
            if not rows:
                raise ValueError("Return Excel has no valid Flipkart return rows.")
            return rows

        missing = [c for c in REQUIRED_RETURN_COLUMNS if c not in headers]
        if missing:
            raise ValueError(f"Return Excel missing required columns: {missing}")
        rows = []
        for row_vals in values_iter:
            row = {}
            for i, h in enumerate(headers):
                row[h] = _normalize_text(row_vals[i] if i < len(row_vals) else "")
            if not _normalize_text(row.get("Suborder Number", "")):
                continue
            rows.append(row)
        if not rows:
            raise ValueError("Return Excel has no valid return rows.")
        return rows
    finally:
        wb.close()


def _order_id_match_keys(value: object) -> list[str]:
    """Return the candidate lookup keys for an Order_id / Suborder Number.

    Flipkart order IDs surface as ``OD...`` on shipping labels but ``OI:...``
    inside the seller returns report, so we index both the normalized value and
    the trailing numeric payload to maximize matching coverage without coupling
    the Meesho suborder path to Flipkart prefixes.
    """
    text = _normalize_suborder(value)
    if not text:
        return []
    keys = [text]
    if text.startswith("od") and text[2:]:
        keys.append(text[2:])
    digits = "".join(ch for ch in text if ch.isdigit())
    if digits and digits not in keys:
        keys.append(digits)
    return keys


def analyze_returns_against_orders(
    *,
    orders_csv_path: str | Path,
    returns_excel_path: str | Path,
    output_dir: str | Path,
) -> tuple[str, dict]:
    orders = _read_orders_csv(orders_csv_path)
    returns = _read_returns_excel(returns_excel_path)
    order_by_suborder: dict[str, dict] = {}
    for row in orders:
        for key in _order_id_match_keys(row.get("Order_id", "")):
            if key and key not in order_by_suborder:
                order_by_suborder[key] = row

    out_rows: list[dict] = []
    matched = 0
    unmatched = 0
    for ret in returns:
        suborder = _normalize_text(ret.get("Suborder Number", ""))
        order: dict | None = None
        for key in _order_id_match_keys(suborder):
            candidate = order_by_suborder.get(key)
            if candidate:
                order = candidate
                break
        if order:
            matched += 1
        else:
            unmatched += 1
            order = {}
        out_rows.append(
            {
                "Suborder Number": suborder,
                "match_status": "matched" if order else "unmatched",
                "Name": _normalize_text(order.get("Name", "")),
                "Pincode": _normalize_text(order.get("Pincode", "")),
                "Sku": _normalize_text(order.get("Sku", "")),
                "Payment_Mode": _normalize_text(order.get("Payment_Mode", "")),
                "Courier_Partner": _normalize_text(order.get("Courier_Partner", "")),
                "Courier_trans_id": _normalize_text(order.get("Courier_trans_id", "")),
                "AWB Number": _normalize_text(ret.get("AWB Number", "")),
                "Type of Return": _normalize_text(ret.get("Type of Return", "")),
                "Sub Type": _normalize_text(ret.get("Sub Type", "")),
                "Status": _normalize_text(ret.get("Status", "")),
                "Return Reason": _normalize_text(ret.get("Return Reason", "")),
                "Detailed Return Reason": _normalize_text(ret.get("Detailed Return Reason", "")),
            }
        )

    out_path = Path(output_dir) / "return-analysis.csv"
    fieldnames = [
        "Suborder Number",
        "match_status",
        "Name",
        "Pincode",
        "Sku",
        "Payment_Mode",
        "Courier_Partner",
        "Courier_trans_id",
        "AWB Number",
        "Type of Return",
        "Sub Type",
        "Status",
        "Return Reason",
        "Detailed Return Reason",
    ]
    with out_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(out_rows)

    summary = {
        "total_order_rows": len(orders),
        "total_returns": len(returns),
        "matched_returns": matched,
        "unmatched_returns": unmatched,
        "match_rate_pct": round((matched / max(1, len(returns))) * 100.0, 2),
    }
    return str(out_path), summary

